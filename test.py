# -*- coding: utf-8 -*-
import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.merge_cells('A1:m1')
# sheet.merge_cells(start_row=1,start_column=1,end_row=4,end_column=1)
sheet['A1'] = 'DBA'
sheet.merge_cells('N1:U1')
sheet['N1'] = '开发同事'
sheet.merge_cells('v1:aa1')
sheet['v1'] = 'DBA组员'
sheet.merge_cells('ab1:ad1')
sheet['ab1'] = '逻辑读'
sheet.merge_cells('ae1:ag1')
sheet['ae1'] = '耗时'
sheet.merge_cells('ah1:aj1')
sheet['ah1'] = 'COSTS'

list  = [['日期', '数据库ip', '应用系统', '开发中心', '开发组', '接口人', '执行用户', 'SQL_ID', 'SQL文本', '问题描述', '逻辑读', '耗时', 'COSTS'],
            ['优化方式', '优化方案', '开发负责人', '处理人', '处理时间', '预计投产时间', '实际投产时间', '备注'],
            ['确认人', '确认时间', '确认结果', '首次通过', '确认意见', '备注'], ['优化前', '优化后', '提升(倍)'], ['优化前', '优化后', '提升(倍)'],
            ['优化前', '优化后', '提升(倍)'], ['更新时间']]


workbook.save('test1.xlsx')